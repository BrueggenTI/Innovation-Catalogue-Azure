"""
Excel Batch Processor
Processes multiple Excel files containing recipe data and merges them based on specification numbers
"""

import openpyxl
import logging
import re
from collections import defaultdict
from nutriscore_mapper import get_nutriscore_image, extract_nutriscore_from_text

logger = logging.getLogger(__name__)

class ExcelBatchProcessor:
    def __init__(self):
        self.recipes_data = {}
    
    def process_batch_files(self, excel_files):
        """
        Process multiple Excel files and extract recipe data
        
        Args:
            excel_files: List of tuples (file_path, file_type, filename)
            
        Returns:
            dict: Dictionary with specification numbers as keys and recipe data as values
        """
        ingredients_data = {}
        nutritional_data = {}
        
        for file_path, file_type, filename in excel_files:
            if file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                # Determine file type based on content
                if self._is_ingredients_file(file_path):
                    ingredients_data = self._parse_ingredients_excel(file_path)
                elif self._is_nutritional_file(file_path):
                    nutritional_data = self._parse_nutritional_excel(file_path)
        
        # Merge data based on specification numbers
        merged_recipes = self._merge_recipe_data(ingredients_data, nutritional_data)
        
        return merged_recipes
    
    def _is_ingredients_file(self, file_path):
        """Check if Excel file contains ingredients data"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            
            # Check for "CompositionLevel" or "Ingredient List" in first column
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                if row[0] and 'CompositionLevel' in str(row[0]):
                    return True
                if row[0] and 'Ingredient' in str(row[0]):
                    return True
            
            return False
        except Exception as e:
            logger.error(f"Error checking ingredients file: {e}")
            return False
    
    def _is_nutritional_file(self, file_path):
        """Check if Excel file contains nutritional data"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            
            # Check for nutritional value headers like "Energy", "Fat", "Nutri Score"
            first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            nutritional_keywords = ['Energy', 'Fat', 'Protein', 'Nutri', 'Score']
            
            for cell in first_row:
                if cell and any(keyword in str(cell) for keyword in nutritional_keywords):
                    return True
            
            return False
        except Exception as e:
            logger.error(f"Error checking nutritional file: {e}")
            return False
    
    def _parse_ingredients_excel(self, file_path):
        """
        Parse ingredients Excel file
        Structure: Columns represent different recipes, rows contain ingredients
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            
            recipes = {}
            
            # Get header row (row 1) with specification numbers
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            # Extract specification numbers from header
            spec_columns = {}
            for col_idx, cell_value in enumerate(header_row):
                if cell_value and col_idx > 1:  # Skip first two columns
                    # Extract specification number (7-digit number)
                    spec_match = re.search(r'\b(\d{7})\b', str(cell_value))
                    if spec_match:
                        spec_num = spec_match.group(1)
                        # Store column index for this specification
                        spec_columns[spec_num] = col_idx
                        recipes[spec_num] = {
                            'name': str(cell_value),
                            'specification': spec_num,
                            'ingredients': []
                        }
            
            # Parse ingredients (ONLY from the row where column A contains "Ingredient" onwards)
            # Example: If "Ingredient" is in row 70, extract ingredients starting from row 70
            ingredient_section = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if we've reached the ingredient section
                # Look for "Ingredient" or "Ingredient List" in column A (index 0)
                cell_value = str(row[0]).strip().lower() if row[0] else ""
                if cell_value in ['ingredient', 'ingredient list']:
                    ingredient_section = True
                    logger.info(f"Found '{row[0]}' marker in column A - starting ingredient extraction from this row")
                    # Don't continue - process this row too if it has data in column B
                
                # Stop extraction if we hit a new section header in column A (but column B is empty)
                # BUT don't stop at the "Ingredient" or "Ingredient List" header itself
                if ingredient_section and row[0] and not row[1]:
                    cell_val_lower = str(row[0]).strip().lower()
                    if cell_val_lower not in ['ingredient', 'ingredient list']:
                        # This is a new section header, stop ingredient extraction
                        logger.info(f"Found new section '{row[0]}' - stopping ingredient extraction")
                        break
                
                # Only extract ingredients AFTER we've found the ingredient section marker
                if ingredient_section and row[1]:  # Row has ingredient data
                    ingredient_spec = row[1]  # Column B has ingredient specification
                    ingredient_name_match = re.search(r'\d+\s+(.+)', str(ingredient_spec))
                    ingredient_name = ingredient_name_match.group(1) if ingredient_name_match else str(ingredient_spec)
                    
                    # Check each recipe column for percentage
                    for spec_num, col_idx in spec_columns.items():
                        if col_idx < len(row) and row[col_idx] and row[col_idx] != '-':
                            percentage = row[col_idx]
                            
                            # Filter out invalid ingredients:
                            # - Values > 100 without % symbol are not ingredients
                            # - Only accept numeric values or values with %
                            if isinstance(percentage, (int, float)):
                                percentage_value = float(percentage)
                                # Exclude values > 100 (these are final products with process loss)
                                if percentage_value <= 100:
                                    recipes[spec_num]['ingredients'].append({
                                        'name': ingredient_name,
                                        'percentage': percentage_value,
                                        'specification': ingredient_spec
                                    })
                            elif isinstance(percentage, str):
                                # Handle percentage strings like "50%" or "12,5%"
                                percentage_clean = percentage.replace(',', '.').replace('%', '').strip()
                                try:
                                    percentage_value = float(percentage_clean)
                                    if percentage_value <= 100:
                                        recipes[spec_num]['ingredients'].append({
                                            'name': ingredient_name,
                                            'percentage': percentage_value,
                                            'specification': ingredient_spec
                                        })
                                except ValueError:
                                    # Skip non-numeric values
                                    pass
            
            logger.info(f"Parsed {len(recipes)} recipes from ingredients file")
            return recipes
            
        except Exception as e:
            logger.error(f"Error parsing ingredients Excel: {e}")
            return {}
    
    def _parse_nutritional_excel(self, file_path):
        """
        Parse nutritional values Excel file
        Each row represents a recipe with its nutritional data
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            
            recipes = {}
            
            # Get header row to identify columns
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            # Map column names to indices
            col_map = {}
            for idx, header in enumerate(header_row):
                if header:
                    col_map[str(header).lower().strip()] = idx
            
            # Parse each row (skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                # Extract specification number from first column
                spec_match = re.search(r'\b(\d{7})\b', str(row[0]))
                if not spec_match:
                    continue
                
                spec_num = spec_match.group(1)
                
                # Extract nutritional data
                # Energy is typically in columns 6 (kJ) and 7 (kcal)
                energy_kj = self._safe_float(row[6]) if len(row) > 6 else 0
                energy_kcal = self._safe_float(row[7]) if len(row) > 7 else 0
                
                recipes[spec_num] = {
                    'name': row[1] if len(row) > 1 else '',
                    'specification': spec_num,
                    'full_specification': row[2] if len(row) > 2 else '',
                    'nutritional_info': {
                        'energy_kj': energy_kj,
                        'energy_kcal': energy_kcal,
                        'energy': energy_kcal,  # Keep for backward compatibility
                        'fat': self._safe_float(row[9]) if len(row) > 9 else 0,
                        'saturated_fat': self._safe_float(row[10]) if len(row) > 10 else 0,
                        'carbohydrates': self._safe_float(row[11]) if len(row) > 11 else 0,
                        'sugars': self._safe_float(row[12]) if len(row) > 12 else 0,
                        'fiber': self._safe_float(row[13]) if len(row) > 13 else 0,
                        'protein': self._safe_float(row[14]) if len(row) > 14 else 0,
                        'salt': self._safe_float(row[15]) if len(row) > 15 else 0
                    },
                    'nutri_score': row[16] if len(row) > 16 else None,
                    'nutri_score_image': get_nutriscore_image(row[16]) if len(row) > 16 else None
                }
            
            logger.info(f"Parsed {len(recipes)} recipes from nutritional file")
            return recipes
            
        except Exception as e:
            logger.error(f"Error parsing nutritional Excel: {e}")
            return {}
    
    def _safe_float(self, value):
        """Safely convert value to float"""
        try:
            if value is None or value == '':
                return 0
            # Handle German decimal format (comma instead of dot)
            if isinstance(value, str):
                value = value.replace(',', '.')
            return float(value)
        except:
            return 0
    
    def _merge_recipe_data(self, ingredients_data, nutritional_data):
        """
        Merge ingredients and nutritional data based on specification numbers
        """
        merged = {}
        
        # Get all unique specification numbers
        all_specs = set(ingredients_data.keys()) | set(nutritional_data.keys())
        
        for spec_num in all_specs:
            recipe = {
                'specification': spec_num,
                'name': '',
                'ingredients': [],
                'nutritional_info': {},
                'nutri_score': None,
                'nutri_score_image': None
            }
            
            # Merge ingredients data
            if spec_num in ingredients_data:
                recipe['ingredients'] = ingredients_data[spec_num].get('ingredients', [])
                if not recipe['name']:
                    recipe['name'] = ingredients_data[spec_num].get('name', '')
            
            # Merge nutritional data
            if spec_num in nutritional_data:
                recipe['nutritional_info'] = nutritional_data[spec_num].get('nutritional_info', {})
                recipe['nutri_score'] = nutritional_data[spec_num].get('nutri_score')
                recipe['nutri_score_image'] = nutritional_data[spec_num].get('nutri_score_image')
                if not recipe['name']:
                    recipe['name'] = nutritional_data[spec_num].get('name', '')
            
            merged[spec_num] = recipe
        
        logger.info(f"Merged {len(merged)} recipes total")
        return merged

    def match_images_to_recipes(self, recipes_data, uploaded_images):
        """
        Match uploaded product images to recipes based on specification number in filename
        
        Args:
            recipes_data: Dictionary of recipe data with spec numbers as keys
            uploaded_images: List of tuples (file_path, filename)
            
        Returns:
            dict: Updated recipes_data with matched images
        """
        import shutil
        import os
        
        for spec_num, recipe in recipes_data.items():
            # Look for images with this spec number in filename
            for image_path, filename in uploaded_images:
                if spec_num in filename:
                    # Copy image to static folder for web access
                    try:
                        static_dir = os.path.join('static', 'images', 'recipes')
                        os.makedirs(static_dir, exist_ok=True)
                        
                        # Create unique filename
                        import time
                        unique_filename = f"batch_{spec_num}_{int(time.time())}_{filename}"
                        static_path = os.path.join(static_dir, unique_filename)
                        
                        # Copy file to static directory
                        shutil.copy2(image_path, static_path)
                        
                        # Store web-accessible URL
                        recipe['image_path'] = f"/static/images/recipes/{unique_filename}"
                        recipe['image_filename'] = filename
                        logger.info(f"Matched and saved image {filename} to recipe {spec_num} at {static_path}")
                        break
                    except Exception as e:
                        logger.error(f"Error copying image for recipe {spec_num}: {e}")
        
        return recipes_data
