
import openpyxl
from openpyxl.styles import Font, Alignment

def create_test_excel_file():
    workbook = openpyxl.Workbook()

    # Sheet 1: Ingredients
    ingredients_sheet = workbook.active
    ingredients_sheet.title = "Ingredients"

    # Header
    ingredients_sheet.merge_cells('A1:A2')
    ingredients_sheet['A1'] = 'Specification number'
    ingredients_sheet.merge_cells('B1:B2')
    ingredients_sheet['B1'] = 'Ingredient'

    ingredients_sheet['C1'] = 'Approved Recipe'
    ingredients_sheet.merge_cells('C1:D1')
    ingredients_sheet['C2'] = '10001'
    ingredients_sheet.merge_cells('C2:D2')

    ingredients_sheet['E1'] = 'Unapproved Recipe'
    ingredients_sheet.merge_cells('E1:F1')
    ingredients_sheet['E2'] = '10002'
    ingredients_sheet.merge_cells('E2:F2')

    # Merged 'Ingredients' cell - mimicking the real file structure
    ingredients_sheet.merge_cells('A3:F3')
    ingredients_cell = ingredients_sheet['A3']
    ingredients_cell.value = 'Ingredients'
    ingredients_cell.font = Font(bold=True)
    ingredients_cell.alignment = Alignment(horizontal='center')

    # Data for Approved Recipe
    ingredients_sheet['A4'] = '12345'
    ingredients_sheet['B4'] = 'Sugar'
    ingredients_sheet['C4'] = 50

    ingredients_sheet['A5'] = '54321'
    ingredients_sheet['B5'] = 'Flour'
    ingredients_sheet['C5'] = 50

    # Data for Unapproved Recipe
    ingredients_sheet['A6'] = '98765'
    ingredients_sheet['B6'] = 'Approved Ingredient'
    ingredients_sheet['E6'] = 30

    ingredients_sheet['A7'] = '60001'
    ingredients_sheet['B7'] = 'Test Ingredient 60001'
    ingredients_sheet['E7'] = 70


    # Sheet 2: Nutritional Information
    nutritional_sheet = workbook.create_sheet(title="Nutritional Information")

    # Header
    nutritional_sheet['A1'] = 'Specification number'
    nutritional_sheet['B1'] = 'Nutrient'
    nutritional_sheet['C1'] = 'Approved Recipe'
    nutritional_sheet['D1'] = '10001'
    nutritional_sheet['E1'] = 'Unapproved Recipe'
    nutritional_sheet['F1'] = '10002'

    # Data
    nutritional_sheet['A2'] = '10001'
    nutritional_sheet['B2'] = 'Energy (kcal)'
    nutritional_sheet['C2'] = 400

    nutritional_sheet['A3'] = '10002'
    nutritional_sheet['B3'] = 'Energy (kcal)'
    nutritional_sheet['E3'] = 500

    workbook.save("Mappe1.xlsx")

if __name__ == "__main__":
    create_test_excel_file()
